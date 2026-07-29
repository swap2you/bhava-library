"""Technical metadata extraction with graceful optional-deps degradation."""

from __future__ import annotations

import importlib.util
import json
import stat
import zipfile
from collections import Counter
from datetime import date, datetime
from pathlib import Path
from typing import Any

from bhava_library.config import Settings
from bhava_library.curation.audit import audited_curation_command
from bhava_library.infrastructure.database import Database, utc_now

EXTRACTOR_VERSION = "enrich-v2.0"
EXTRACTOR_MODULES = {
    "pypdf": "pypdf",
    "mutagen": "mutagen",
    "Pillow": "PIL",
    "python-docx": "docx",
    "python-pptx": "pptx",
    "openpyxl": "openpyxl",
}
ARCHIVE_SUFFIXES = {".zip", ".tar", ".gz", ".tgz", ".7z", ".rar"}
ZIP_SUFFIXES = {".zip"}
QUARANTINE_PARTS = {"quarantine", "quarantined"}


def extractor_availability() -> dict[str, bool]:
    """Report optional extractor availability without importing file parsers."""
    return {
        name: importlib.util.find_spec(module) is not None
        for name, module in EXTRACTOR_MODULES.items()
    }


def _json_value(value: Any) -> Any:
    if isinstance(value, (date, datetime)):
        return value.isoformat()
    if isinstance(value, (str, int, float, bool)) or value is None:
        return value
    return str(value)


def _guess_mime(path: Path) -> str | None:
    try:
        import filetype

        kind = filetype.guess(str(path))
        return kind.mime if kind else None
    except Exception:  # noqa: BLE001
        return None


def _pdf_metadata(path: Path) -> dict[str, Any]:
    from pypdf import PdfReader

    reader = PdfReader(str(path))
    info: dict[str, Any] = dict(reader.metadata or {})
    searchable = False
    text_errors: list[str] = []
    for page in reader.pages[: min(10, len(reader.pages))]:
        try:
            if (page.extract_text() or "").strip():
                searchable = True
                break
        except Exception as exc:  # noqa: BLE001
            text_errors.append(f"{type(exc).__name__}: {exc}")
    result = {
        "page_count": len(reader.pages),
        "pdf_info": {str(k): _json_value(v) for k, v in info.items() if v is not None},
        "searchable_text_available": searchable,
        "likely_scanned": bool(reader.pages) and not searchable,
        "extractor": "pypdf",
    }
    if text_errors:
        result["text_extraction_errors"] = text_errors
    return result


def _audio_metadata(path: Path) -> dict[str, Any]:
    from mutagen import File as MutagenFile

    audio = MutagenFile(path)
    if audio is None:
        raise ValueError("mutagen did not recognize the audio file")
    info = getattr(audio, "info", None)
    tags = {
        str(key): _json_value(value) for key, value in (getattr(audio, "tags", None) or {}).items()
    }
    return {
        "duration_seconds": getattr(info, "length", None),
        "codec": type(info).__name__ if info is not None else None,
        "bitrate": getattr(info, "bitrate", None),
        "sample_rate": getattr(info, "sample_rate", None),
        "channels": getattr(info, "channels", None),
        "tags": tags,
        "extractor": "mutagen",
    }


def _core_properties(props: Any) -> dict[str, Any]:
    names = (
        "title",
        "subject",
        "author",
        "keywords",
        "comments",
        "last_modified_by",
        "created",
        "modified",
        "category",
        "content_status",
        "identifier",
        "language",
        "version",
    )
    return {
        name: _json_value(getattr(props, name, None))
        for name in names
        if getattr(props, name, None) is not None
    }


def _office_metadata(path: Path, suffix: str) -> dict[str, Any]:
    if suffix == ".docx":
        from docx import Document

        document = Document(str(path))
        return {
            "document_type": "word",
            "core_properties": _core_properties(document.core_properties),
            "extractor": "python-docx",
        }
    if suffix == ".pptx":
        from pptx import Presentation

        presentation = Presentation(str(path))
        return {
            "document_type": "presentation",
            "slide_count": len(presentation.slides),
            "core_properties": _core_properties(presentation.core_properties),
            "extractor": "python-pptx",
        }
    from openpyxl import load_workbook

    workbook = load_workbook(path, read_only=True, data_only=True, keep_vba=False)
    try:
        return {
            "document_type": "spreadsheet",
            "sheet_count": len(workbook.sheetnames),
            "sheet_names": list(workbook.sheetnames),
            "core_properties": _core_properties(workbook.properties),
            "extractor": "openpyxl",
        }
    finally:
        workbook.close()


def _image_metadata(path: Path) -> dict[str, Any]:
    from PIL import Image

    with Image.open(path) as image:
        return {
            "width": image.width,
            "height": image.height,
            "mode": image.mode,
            "format": image.format,
            "extractor": "Pillow",
        }


def _archive_metadata(path: Path) -> dict[str, Any]:
    if not zipfile.is_zipfile(path):
        raise ValueError("file is not a valid ZIP archive")
    with zipfile.ZipFile(path) as zf:
        members = zf.infolist()
        executable = []
        encrypted = []
        for member in members:
            mode = (member.external_attr >> 16) & 0xFFFF
            if mode and stat.S_ISREG(mode) and mode & 0o111:
                executable.append(member.filename)
            if member.flag_bits & 0x1:
                encrypted.append(member.filename)
        return {
            "archive_type": "zip",
            "member_count": len(members),
            "uncompressed_bytes": sum(member.file_size for member in members),
            "has_encrypted_members": bool(encrypted),
            "encrypted_member_count": len(encrypted),
            "has_executable_members": bool(executable),
            "executable_member_count": len(executable),
            "sample_members": [member.filename for member in members[:50]],
            "extractor": "stdlib-zipfile",
        }


def extract_technical_metadata(path: Path, *, quarantined: bool | None = None) -> dict[str, Any]:
    suffix = path.suffix.lower()
    available = extractor_availability()
    payload: dict[str, Any] = {
        "filename": path.name,
        "extension": suffix,
        "size_bytes": path.stat().st_size if path.exists() else None,
        "mime_guess": _guess_mime(path) if path.exists() else None,
        "extractor_availability": available,
        "extraction_status": "fallback_only",
        "errors": [],
    }
    is_quarantined = (
        quarantined
        if quarantined is not None
        else bool(QUARANTINE_PARTS.intersection(part.lower() for part in path.parts))
    )
    if is_quarantined and suffix in ARCHIVE_SUFFIXES:
        payload["extraction_status"] = "skipped_quarantine"
        return payload
    if not path.exists():
        payload["extraction_status"] = "error"
        payload["errors"].append("FileNotFoundError: file does not exist")
        return payload

    extractor: tuple[str, str, Any] | None = None
    if suffix == ".pdf":
        extractor = ("pdf", "pypdf", _pdf_metadata)
    elif suffix in {".mp3", ".m4a", ".ogg", ".wav", ".flac"}:
        extractor = ("audio", "mutagen", _audio_metadata)
    elif suffix in {".docx", ".pptx", ".xlsx"}:
        package = {".docx": "python-docx", ".pptx": "python-pptx", ".xlsx": "openpyxl"}[suffix]
        extractor = ("office", package, lambda value: _office_metadata(value, suffix))
    elif suffix in {".png", ".jpg", ".jpeg", ".gif", ".webp", ".tif", ".tiff", ".bmp"}:
        extractor = ("image", "Pillow", _image_metadata)
    elif suffix in ZIP_SUFFIXES:
        extractor = ("archive", "stdlib-zipfile", _archive_metadata)

    if extractor is None:
        return payload
    section, package, function = extractor
    if package != "stdlib-zipfile" and not available[package]:
        payload["errors"].append(f"optional extractor unavailable: {package}")
        return payload
    try:
        payload[section] = function(path)
        partial = bool(payload[section].get("text_extraction_errors"))
        payload["extraction_status"] = "partial" if partial else "full"
    except Exception as exc:  # noqa: BLE001
        payload["extraction_status"] = "error"
        payload["errors"].append(f"{type(exc).__name__}: {exc}")
    return payload


def summarize_metadata_coverage(payloads: list[dict[str, Any]]) -> dict[str, Any]:
    statuses = Counter(
        str(payload.get("extraction_status", "fallback_only")) for payload in payloads
    )
    full_or_partial = statuses["full"] + statuses["partial"]
    return {
        "total": len(payloads),
        "status_counts": dict(sorted(statuses.items())),
        "technical_metadata_count": full_or_partial,
        "coverage_percent": round((full_or_partial / len(payloads) * 100), 2) if payloads else 0.0,
    }


def metadata_coverage_report(settings: Settings) -> dict[str, Any]:
    db = Database(settings.catalog_db)
    db.migrate()
    rows = db.execute("SELECT payload_json FROM technical_metadata ORDER BY resource_id")
    payloads = [json.loads(row["payload_json"]) for row in rows]
    report = summarize_metadata_coverage(payloads)
    resource_count = db.execute("SELECT COUNT(*) AS count FROM resources WHERE removed_at IS NULL")[
        0
    ]["count"]
    report["active_resources"] = resource_count
    report["recorded_metadata"] = len(payloads)
    report["record_coverage_percent"] = (
        round(len(payloads) / resource_count * 100, 2) if resource_count else 0.0
    )
    return report


def _sidecar_paths(settings: Settings, resource_id: str) -> tuple[Path, Path]:
    technical = settings.data_dir / "derived" / "technical" / f"{resource_id}.json"
    metadata = settings.data_dir / "derived" / "metadata" / f"{resource_id}.json"
    technical.parent.mkdir(parents=True, exist_ok=True)
    metadata.parent.mkdir(parents=True, exist_ok=True)
    return technical, metadata


def write_sidecars(
    settings: Settings,
    resource_id: str,
    payload: dict[str, Any],
    *,
    catalog_row: dict[str, Any] | None = None,
) -> tuple[Path, Path]:
    technical_path, metadata_path = _sidecar_paths(settings, resource_id)
    envelope = {
        "resource_id": resource_id,
        "extractor_version": EXTRACTOR_VERSION,
        "extracted_at": utc_now(),
        "technical": payload,
    }
    technical_path.write_text(json.dumps(envelope, indent=2), encoding="utf-8")
    meta = {
        "resource_id": resource_id,
        "display_sidecar": str(metadata_path.relative_to(settings.data_dir)),
        "catalog": catalog_row or {},
    }
    metadata_path.write_text(json.dumps(meta, indent=2), encoding="utf-8")
    return technical_path, metadata_path


@audited_curation_command("enrich")
def run_enrich(settings: Settings, *, limit: int | None = None) -> dict[str, int]:
    db = Database(settings.catalog_db)
    db.migrate()
    sql = """
        SELECT r.resource_id, r.title_original, r.media_type, r.media_format,
               lf.relative_path, lf.size_bytes, lf.sha256
        FROM resources r
        JOIN local_files lf ON lf.resource_id = r.resource_id
        WHERE r.removed_at IS NULL
        ORDER BY r.resource_id
    """
    params: tuple[()] | tuple[int] = ()
    if limit is not None:
        sql += " LIMIT ?"
        params = (limit,)
    rows = db.execute(sql, params)
    enriched = 0
    with db.session() as conn:
        for row in rows:
            rel = row["relative_path"]
            disk_path = settings.repo_root / rel if not Path(rel).is_absolute() else Path(rel)
            if not disk_path.exists():
                disk_path = settings.data_dir.parent / rel
            if not disk_path.exists() and rel.startswith("data/"):
                disk_path = settings.repo_root / rel
            payload = extract_technical_metadata(disk_path)
            write_sidecars(
                settings,
                row["resource_id"],
                payload,
                catalog_row={
                    "title_original": row["title_original"],
                    "relative_path": rel,
                    "size_bytes": row["size_bytes"],
                    "sha256": row["sha256"],
                },
            )
            conn.execute(
                """
                INSERT INTO technical_metadata(resource_id, payload_json, extracted_at, extractor_version)
                VALUES (?, ?, ?, ?)
                ON CONFLICT(resource_id) DO UPDATE SET
                  payload_json = excluded.payload_json,
                  extracted_at = excluded.extracted_at,
                  extractor_version = excluded.extractor_version
                """,
                (
                    row["resource_id"],
                    json.dumps(payload),
                    utc_now(),
                    EXTRACTOR_VERSION,
                ),
            )
            enriched += 1
    coverage = metadata_coverage_report(settings)
    return {
        "candidates": len(rows),
        "enriched": enriched,
        "full": coverage["status_counts"].get("full", 0),
        "partial": coverage["status_counts"].get("partial", 0),
        "fallback_only": coverage["status_counts"].get("fallback_only", 0),
        "skipped_quarantine": coverage["status_counts"].get("skipped_quarantine", 0),
        "errors": coverage["status_counts"].get("error", 0),
    }
